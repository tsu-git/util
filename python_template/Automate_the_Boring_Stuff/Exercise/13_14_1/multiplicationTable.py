"""multiplicationTable

    掛け算の表を作成する
"""

import openpyxl, sys, argparse
from openpyxl.styles import Font

ROW_INIT = 1
COL_INIT = 1

# 引数から掛け算を行う数を受け取る
parser = argparse.ArgumentParser(
            prog='multiplicationTable.py',
            description='make a multiplication table')

parser.add_argument('number')
args = parser.parse_args()
N = int(args.number) + 1

# ワークブックオブジェクトを生成する
wb = openpyxl.Workbook()
sht = wb.active

# ワークシートに掛け算表を出力する
font_obj = Font(bold=True)
#  横のヘッダ
for i in range(1, N):
    sht.cell(row=ROW_INIT, column=COL_INIT+i).value = i
    # フォントをボールドにする
    sht.cell(row=ROW_INIT, column=COL_INIT+i).font = font_obj

#  縦のヘッダ
for j in range(1, N):
    sht.cell(row=ROW_INIT+j, column=COL_INIT).value = j
    # フォントをボールドにする
    sht.cell(row=ROW_INIT+j, column=COL_INIT).font = font_obj

#  掛け算結果
for j in range(1, N):
    for i in range(1, N):
        sht.cell(row=ROW_INIT+j, column=COL_INIT+i).value = i * j

wb.save('multiplicationTable.xlsx')

sys.exit()
