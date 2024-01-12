"""translate_excel_to_text

    スプレッドシートからテキストファイルに変換する。
    列ごとにファイルを分ける。
"""
import openpyxl, sys, argparse, logging
from pathlib import Path

# 引数を取得する
parser = argparse.ArgumentParser(
            prog='translate_excel_to_text',
            description='translate from a Excel file into text files'
        )
parser.add_argument('input_excel')
parser.add_argument('-o', '--out_filestem', default='out',
                    help='output filename without extension')
parser.add_argument('-d', '--debug', action='store_true')
args = parser.parse_args()

# ログ出力の準備をする
LOG_FMT = "%(asctime)s: %(levelname)s: %(message)s"
logging.basicConfig(format=LOG_FMT, level=logging.DEBUG)
if args.debug is False:
    logging.disable(logging.WARNING)
logging.debug(f'start {sys.argv[0]}')

# 入力ファイルの検査をする
if Path(args.input_excel).is_file() is False:
    logging.error(f'{args.input_excel} is not a file')
    sys.exit(False)

ext = (args.input_excel).split('.', maxsplit=1)[1] 
if ext != 'xlsx':
    logging.error(f'{args.input_excel} does not have a excel extension')
    sys.exit(False)

# 引数のExcelファイルを読み込む
wb = openpyxl.load_workbook(args.input_excel)
sht = wb.active

# ファイルを一列ずつ読み込む
for i in range(sht.max_column):
    with open(f"{args.out_filestem}_{i}.txt", "w") as f:
        for j in range(sht.max_row):
            f.write(f"{sht.cell(row=j+1, column=i+1).value}")
    

sys.exit(True)
