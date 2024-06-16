"""upd_delivery_list.py
    納品一覧の項目のいくつかを、手入力値に基づいて計算する

        対応済みの項目
        - カラム区分

"""

import sys
from pathlib import Path
import openpyxl
import re


# ファイル名の形式チェック関数
def examine_filename_format(filename: str) -> bool:
    if (matched := re.search(r'^[WM][LG][1-3][23]-', filename)):
        return(True)
    else:
        return(False)


# 使用量、計器数の判定関数
def find_out_usage_and_meter(filename: str) -> str:
    kind = ""

    if filename[0] == "W":
        kind = "使用量"
    elif filename[0] == "M":
        kind = "計器数"

    return(kind)


# 電圧判定関数
def find_out_voltage(filename: str) -> str:
    kind = ""

    if filename[2] == "1":
        kind = "低圧"
    elif filename[2] == "2":
        kind = "高圧"
    elif filename[2] == "3":
        kind = "特高"

    return(kind)

def find_out_column_kbn(filename: str) -> str:
    kind = ""

    if filename[:2] == "WL":
        if filename[2] == "1":
            kind = "1"
        else:
            kind = "2"
    elif filename[:2] == "WG":
        kind = "3"
    elif filename[:2] == "ML":
        if filename[2] == "1":
            kind = "4"
        else:
            kind = "5"
    elif filename[:2] == "MG":
        if filename[2] == "1":
            kind = "6"
        else:
            kind = "7"
    else:
        kind = "999"

    return(kind)

# 定数定義
HEADER_TARGET = "No."   # ヘッダ行のいずれかの項目
SEARCH_RANGE_ROW_MAX = 50
SEARCH_RANGE_COL_MAX = 30
COL_IDX_FILENAME = 2
COL_IDX_COLUMN_KBN = 3

# 引数取得
if len(sys.argv) < 2:
    print(f"Usage: {sys.argv[0]} delivery_list_file.xlsx")
    sys.exit()

list_file = Path(sys.argv[1])
if list_file.is_file() != True:
    print(f"{list_file} dose not exist")
    print(f"Usage: {sys.argv[0]} delivery_list_file.xlsx")
    sys.exit()

if list_file.suffix != ".xlsx":
    print(f"{list_file} is not a excel file")
    print(f"Usage: {sys.argv[0]} delivery_list_file.xlsx")
    sys.exit()


# 納品一覧ファイルを開く
wb = openpyxl.load_workbook(list_file)

# 対象のシートを選択する
sheet = wb['Sheet1']

# ヘッダ行を探す
for row_num in range(sheet.min_row, SEARCH_RANGE_ROW_MAX):
    for col_num in range(sheet.min_column, SEARCH_RANGE_COL_MAX):
        label = sheet.cell(row=row_num, column=col_num).value
        if label == HEADER_TARGET:
            match_found = True
            break
    if match_found:
        break

header_row = row_num

# TODO: ファイル名から項目を自動計算する
for row_num in range(header_row + 1, sheet.max_row):
    filename = sheet.cell(row=row_num, column=COL_IDX_FILENAME).value

    # ファイル名検査
    if filename == None:
        # ファイル名が空の場合終了
        break
    elif filename[-4:] != ".csv":
        # csvファイル以外は無視
        continue

    print(filename, end=" ")

    # 秘匿化前ファイルの場合はファイル名からキーワードを除去する
    if filename[0:10] == "afterhide_":
        filename = filename[10:]

    if examine_filename_format(filename) != True:
        continue

    # 使用量／計器数
    print(find_out_usage_and_meter(filename), end=" ")
    sheet.cell(row=row_num, column=3).value = find_out_usage_and_meter(filename)

    # 電圧
    print(find_out_voltage(filename), end=" ")
    sheet.cell(row=row_num, column=4).value = find_out_voltage(filename)

    # カラム区分
    print(find_out_column_kbn(filename))
    sheet.cell(row=row_num, column=6).value = find_out_column_kbn(filename)


# 納品一覧ファイルを保存
wb.save(list_file)

# 納品一覧ファイルを閉じる
wb.close()

sys.exit()
